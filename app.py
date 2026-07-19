"""
Conciliador Bancário Inteligente
---------------------------------
App de conciliação bancária (Extrato x Sistema) para escritórios de
contabilidade e empresas de pequeno/médio porte.

Para configurar em produção, crie um arquivo .streamlit/secrets.toml com:

    [usuarios]
    nome_do_cliente = "hash_sha256_da_senha"

Gere o hash com: python -c "import hashlib; print(hashlib.sha256('SUASENHA'.encode()).hexdigest())"

Sem esse arquivo, o app roda em modo demonstração com usuário "demo" / senha "galvao3dias".
"""
import io
import re
import hashlib
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# CONFIGURAÇÃO DA PÁGINA
# =====================================================================
st.set_page_config(
    page_title="Conciliador Bancário Inteligente",
    page_icon="📒",
    layout="wide",
    initial_sidebar_state="expanded",
)

NOME_PRODUTO = "Conciliador Bancário Inteligente"


# =====================================================================
# ESTILO — sistema visual inspirado no razão contábil clássico:
# papel, tinta, verde-livro-caixa e o "carimbo" de conferência.
# =====================================================================
def injetar_estilo():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500;600&display=swap');

        :root {
            --papel: #FAF9F4;
            --tinta: #1F2421;
            --tinta-suave: #5B6660;
            --verde: #2F4B3C;
            --verde-claro: #E7EDE6;
            --verde-hover: #24392D;
            --vermelho: #A8342A;
            --vermelho-claro: #F7E8E6;
            --ambar: #A8703A;
            --ambar-claro: #F5EBDD;
            --linha: #D8D4C6;
        }

        html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
        .stApp { background-color: var(--papel); }
        h1, h2, h3 { font-family: 'Fraunces', serif !important; color: var(--verde) !important; letter-spacing: -0.01em; }

        [data-testid="stMetricValue"], .cb-mono { font-family: 'IBM Plex Mono', monospace !important; font-variant-numeric: tabular-nums; }

        .cb-header {
            background: linear-gradient(135deg, var(--verde) 0%, #1C3226 100%);
            border-radius: 12px; padding: 2.1rem 2.3rem; margin-bottom: 1.4rem;
            position: relative; overflow: hidden;
        }
        .cb-header::after {
            content: ""; position: absolute; top: -45%; right: -8%;
            width: 260px; height: 260px; border: 1px solid rgba(255,255,255,0.12); border-radius: 50%;
        }
        .cb-header-eyebrow {
            font-family: 'IBM Plex Mono', monospace; text-transform: uppercase;
            letter-spacing: 0.12em; font-size: 0.7rem; color: #A9C4B2; margin-bottom: 0.4rem;
        }
        .cb-header h1 { color: #FFFFFF !important; font-size: 1.9rem; margin: 0 0 0.3rem 0; }
        .cb-header p { color: #D9E6DD; margin: 0; font-size: 0.93rem; max-width: 640px; }

        .cb-card {
            background: #FFFFFF; border: 1px solid var(--linha); border-radius: 10px;
            padding: 1.05rem 1.25rem; height: 100%;
        }
        .cb-card-label { font-size: 0.74rem; color: var(--tinta-suave); text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 0.3rem; }
        .cb-card-value { font-family: 'IBM Plex Mono', monospace; font-size: 1.7rem; font-weight: 600; line-height: 1.15; color: var(--tinta); }
        .cb-card-sub { font-size: 0.78rem; color: var(--tinta-suave); margin-top: 0.25rem; }
        .cb-card--verde { border-left: 3px solid var(--verde); }
        .cb-card--vermelho { border-left: 3px solid var(--vermelho); }
        .cb-card--ambar { border-left: 3px solid var(--ambar); }
        .cb-card--neutro { border-left: 3px solid var(--linha); }

        .cb-selo-wrap {
            display: flex; align-items: center; gap: 1.5rem; background: #FFFFFF;
            border: 1px solid var(--linha); border-radius: 10px; padding: 1.2rem 1.5rem; height: 100%;
        }
        .cb-selo { width: 116px; height: 116px; border-radius: 50%; display: flex; align-items: center; justify-content: center; flex-shrink: 0; }
        .cb-selo-inner {
            width: 90px; height: 90px; background: var(--papel); border: 1.5px dashed rgba(47,75,60,0.35);
            border-radius: 50%; display: flex; flex-direction: column; align-items: center; justify-content: center;
        }
        .cb-selo-pct { font-family: 'IBM Plex Mono', monospace; font-size: 1.5rem; font-weight: 600; color: var(--verde); }
        .cb-selo-label { font-size: 0.6rem; text-transform: uppercase; letter-spacing: 0.08em; color: var(--tinta-suave); }
        .cb-selo-titulo { font-family: 'Fraunces', serif; font-size: 1.05rem; color: var(--verde); font-weight: 600; }
        .cb-selo-desc { color: var(--tinta-suave); font-size: 0.85rem; margin-top: 0.15rem; }

        [data-testid="stSidebar"] { background-color: var(--verde); }
        [data-testid="stSidebar"] * { color: #EAF0EC !important; }
        [data-testid="stSidebar"] input { color: var(--tinta) !important; }
        [data-testid="stSidebar"] [data-baseweb="select"] * { color: var(--tinta) !important; }

        .stButton > button, .stDownloadButton > button {
            background-color: var(--verde); color: #FFFFFF !important; border: none; border-radius: 8px;
            font-weight: 600; padding: 0.5rem 1.1rem; transition: background-color 0.15s ease;
        }
        .stButton > button:hover, .stDownloadButton > button:hover { background-color: var(--verde-hover); }

        .cb-divisor { border: none; border-top: 1px solid var(--linha); margin: 1.5rem 0; }
        .cb-aviso-demo {
            background: var(--ambar-claro); border: 1px solid var(--ambar); color: #6B4A22 !important;
            border-radius: 8px; padding: 0.6rem 0.9rem; font-size: 0.82rem; margin-bottom: 1rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


# =====================================================================
# AUTENTICAÇÃO
# Compara hashes (nunca texto puro) e busca credenciais em st.secrets,
# nunca no código-fonte. Mantém um fallback de demonstração, visível
# como tal, para o app não quebrar antes da configuração em produção.
# =====================================================================
def hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode("utf-8")).hexdigest()


USUARIOS_DEMO = {"demo": hash_senha("galvao3dias")}


def carregar_usuarios():
    try:
        usuarios_secrets = dict(st.secrets["usuarios"])
        if usuarios_secrets:
            return usuarios_secrets, True
    except Exception:
        pass
    return USUARIOS_DEMO, False


def tela_login():
    usuarios, producao = carregar_usuarios()

    _, col_meio, _ = st.columns([1, 1.2, 1])
    with col_meio:
        st.markdown(
            f"""<div class="cb-header" style="margin-top:2.5rem;">
                <div class="cb-header-eyebrow">Acesso restrito</div>
                <h1>📒 {NOME_PRODUTO}</h1>
                <p>Cruze o extrato bancário com o sistema em segundos e feche o mês com confiança.</p>
            </div>""",
            unsafe_allow_html=True,
        )

        if not producao:
            st.markdown(
                '<div class="cb-aviso-demo">⚠️ Rodando em modo demonstração '
                '(usuário <b>demo</b> / senha <b>galvao3dias</b>). '
                'Configure <code>st.secrets["usuarios"]</code> antes de usar com clientes reais.</div>',
                unsafe_allow_html=True,
            )

        with st.form("form_login"):
            usuario = st.text_input("Usuário")
            senha = st.text_input("Senha", type="password")
            entrar = st.form_submit_button("Entrar", use_container_width=True)

        if entrar:
            if usuario in usuarios and usuarios[usuario] == hash_senha(senha):
                st.session_state.autenticado = True
                st.session_state.usuario = usuario
                st.rerun()
            else:
                st.error("❌ Usuário ou senha incorretos.")


# =====================================================================
# MOTOR DE CONCILIAÇÃO
# (lógica testada isoladamente antes de entrar no app — ver notas no README)
# =====================================================================
def limpar_valor_monetario(valor):
    """Converte texto/número para float, tratando formato BR (1.234,56),
    internacional (1,234.56), prefixo R$, parênteses para negativo e
    milhar com ponto sem casas decimais (2.500 -> 2500)."""
    if pd.isna(valor):
        return np.nan
    if isinstance(valor, (int, float, np.integer, np.floating)):
        return float(valor)

    texto = str(valor).strip()
    if texto == "":
        return np.nan

    negativo = False
    if texto.startswith("(") and texto.endswith(")"):
        negativo = True
        texto = texto[1:-1].strip()

    texto = re.sub(r"[Rr]\$\s*", "", texto)
    texto = texto.strip().replace(" ", "")

    if texto.startswith("-"):
        negativo = True
        texto = texto[1:]
    elif texto.startswith("+"):
        texto = texto[1:]

    if texto == "":
        return np.nan

    tem_virgula = "," in texto
    tem_ponto = "." in texto

    if tem_virgula and tem_ponto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif tem_virgula:
        texto = texto.replace(",", ".")
    elif tem_ponto:
        partes = texto.split(".")
        if len(partes) > 2 or (len(partes) == 2 and len(partes[1]) == 3):
            texto = texto.replace(".", "")

    try:
        numero = float(texto)
    except ValueError:
        return np.nan
    return -numero if negativo else numero


def limpar_coluna_valores(serie):
    return serie.apply(limpar_valor_monetario)


def detectar_colunas_candidatas(colunas, palavras_chave):
    candidatas = []
    for col in colunas:
        col_norm = str(col).strip().lower()
        if any(p in col_norm for p in palavras_chave):
            candidatas.append(col)
    return candidatas


def preparar_dataframe(df_raw, col_data, col_valor, usar_valor_absoluto=False):
    """Padroniza um DataFrame bruto em colunas 'Data' e 'Valor' limpas,
    preservando as demais colunas originais (exceto as duas usadas como fonte,
    para não gerar nomes de coluna duplicados)."""
    df = df_raw.copy()
    valor_serie = limpar_coluna_valores(df[col_valor])
    if usar_valor_absoluto:
        valor_serie = valor_serie.abs()
    data_serie = pd.to_datetime(df[col_data], errors="coerce", dayfirst=True)

    outras_colunas = [c for c in df.columns if c not in (col_data, col_valor)]
    df_final = df[outras_colunas].copy()
    df_final.insert(0, "Valor", valor_serie.values)
    df_final.insert(0, "Data", data_serie.values)

    linhas_invalidas = int((df_final["Data"].isna() | df_final["Valor"].isna()).sum())
    df_valido = df_final[df_final["Data"].notna() & df_final["Valor"].notna()].copy()
    return df_valido, linhas_invalidas


def executar_conciliacao(df_banco, df_sistema, tolerancia_dias=0):
    banco = df_banco.reset_index(drop=True).copy()
    sistema = df_sistema.reset_index(drop=True).copy()

    banco["_occ"] = banco.groupby(["Data", "Valor"]).cumcount()
    sistema["_occ"] = sistema.groupby(["Data", "Valor"]).cumcount()

    colunas_banco = [c for c in banco.columns if c not in ("Data", "Valor", "_occ")]
    colunas_sistema = [c for c in sistema.columns if c not in ("Data", "Valor", "_occ")]

    banco_r = banco.rename(columns={c: f"{c}_Banco" for c in colunas_banco})
    sistema_r = sistema.rename(columns={c: f"{c}_Sistema" for c in colunas_sistema})
    colunas_banco_r = [f"{c}_Banco" for c in colunas_banco]
    colunas_sistema_r = [f"{c}_Sistema" for c in colunas_sistema]

    merge_exato = pd.merge(
        banco_r, sistema_r, on=["Data", "Valor", "_occ"], how="outer", indicator=True
    )

    exatos = merge_exato[merge_exato["_merge"] == "both"].copy()
    exatos["Data_Banco"] = exatos["Data"]
    exatos["Data_Sistema"] = exatos["Data"]
    exatos["Valor_Banco"] = exatos["Valor"]
    exatos["Valor_Sistema"] = exatos["Valor"]
    exatos["Tipo_de_Match"] = "Exato"
    exatos["Diferenca_Dias"] = 0
    exatos = exatos.drop(columns=["_merge", "_occ", "Data", "Valor"])

    apenas_banco = merge_exato[merge_exato["_merge"] == "left_only"].drop(columns=["_merge", "_occ"]).copy()
    apenas_sistema = merge_exato[merge_exato["_merge"] == "right_only"].drop(columns=["_merge", "_occ"]).copy()
    apenas_banco = apenas_banco.drop(columns=[c for c in colunas_sistema_r if c in apenas_banco.columns])
    apenas_sistema = apenas_sistema.drop(columns=[c for c in colunas_banco_r if c in apenas_sistema.columns])

    aproximados = pd.DataFrame()
    if tolerancia_dias > 0 and not apenas_banco.empty and not apenas_sistema.empty:
        pool_sistema = apenas_sistema.copy()
        pool_sistema["_usado"] = False
        linhas_aprox = []
        indices_banco_usados = []

        for idx_b, linha_b in apenas_banco.iterrows():
            candidatos = pool_sistema[(pool_sistema["Valor"] == linha_b["Valor"]) & (~pool_sistema["_usado"])].copy()
            if candidatos.empty:
                continue
            candidatos["_dif_dias"] = (pd.to_datetime(candidatos["Data"]) - pd.to_datetime(linha_b["Data"])).abs().dt.days
            candidatos = candidatos[candidatos["_dif_dias"] <= tolerancia_dias].sort_values("_dif_dias")
            if candidatos.empty:
                continue
            melhor_idx = candidatos.index[0]
            melhor = pool_sistema.loc[melhor_idx]
            dif_dias = int(candidatos.loc[melhor_idx, "_dif_dias"])

            linha_combinada = {}
            for col in colunas_banco_r:
                linha_combinada[col] = linha_b.get(col, np.nan)
            for col in colunas_sistema_r:
                linha_combinada[col] = melhor.get(col, np.nan)
            linha_combinada["Data_Banco"] = linha_b["Data"]
            linha_combinada["Valor_Banco"] = linha_b["Valor"]
            linha_combinada["Data_Sistema"] = melhor["Data"]
            linha_combinada["Valor_Sistema"] = melhor["Valor"]
            linha_combinada["Tipo_de_Match"] = f"Aproximado ({dif_dias} dia(s))"
            linha_combinada["Diferenca_Dias"] = dif_dias

            linhas_aprox.append(linha_combinada)
            pool_sistema.loc[melhor_idx, "_usado"] = True
            indices_banco_usados.append(idx_b)

        if linhas_aprox:
            aproximados = pd.DataFrame(linhas_aprox)
            apenas_banco = apenas_banco.drop(index=indices_banco_usados)
            usados_idx = pool_sistema[pool_sistema["_usado"]].index
            apenas_sistema = apenas_sistema.drop(index=usados_idx)

    conciliados = pd.concat([exatos, aproximados], ignore_index=True) if not aproximados.empty else exatos.reset_index(drop=True)

    apenas_banco = apenas_banco.rename(columns={"Data": "Data_Banco", "Valor": "Valor_Banco"})
    apenas_sistema = apenas_sistema.rename(columns={"Data": "Data_Sistema", "Valor": "Valor_Sistema"})
    apenas_banco["Status"] = "Apenas no Banco (não lançado no Sistema)"
    apenas_sistema["Status"] = "Apenas no Sistema (não consta no Extrato Bancário)"

    divergencias = pd.concat([apenas_banco, apenas_sistema], ignore_index=True, sort=False)
    return conciliados, divergencias


# =====================================================================
# EXPORTAÇÃO — Excel estilizado (fonte profissional, cabeçalho colorido,
# aba de resumo executivo) pronto para ser enviado ao cliente.
# =====================================================================
def _estilizar_planilha(ws, df, cor_cabecalho="2F4B3C"):
    fill = PatternFill(start_color=cor_cabecalho, end_color=cor_cabecalho, fill_type="solid")
    fonte_cabecalho = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    fonte_corpo = Font(name="Calibri", size=10.5)
    borda_fina = Border(bottom=Side(style="thin", color="D9D9D9"))

    for col_num in range(1, len(df.columns) + 1):
        celula = ws.cell(row=1, column=col_num)
        celula.fill = fill
        celula.font = fonte_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_num in range(2, len(df) + 2):
        for col_num in range(1, len(df.columns) + 1):
            celula = ws.cell(row=row_num, column=col_num)
            celula.font = fonte_corpo
            celula.border = borda_fina

    for col_num, nome_col in enumerate(df.columns, start=1):
        if len(df):
            maior = max([len(str(nome_col))] + [len(str(v)) if pd.notna(v) else 0 for v in df[nome_col]])
        else:
            maior = len(str(nome_col))
        ws.column_dimensions[get_column_letter(col_num)].width = min(max(maior + 3, 12), 42)

    if len(df) > 0:
        ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24


def gerar_excel_conciliacao(conciliados, divergencias, resumo, nome_empresa=""):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        linhas_resumo = [
            ["Relatório de Conciliação Bancária"],
            [nome_empresa] if nome_empresa else [""],
            [f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"],
            [],
            ["Indicador", "Valor"],
            ["Total de lançamentos no extrato bancário", resumo["total_banco"]],
            ["Total de lançamentos no sistema", resumo["total_sistema"]],
            ["Lançamentos conciliados", resumo["total_conciliados"]],
            ["Divergências", resumo["total_divergencias"]],
            ["Índice de conciliação", f'{resumo["percentual"]:.1f}%'],
        ]
        pd.DataFrame(linhas_resumo).to_excel(writer, sheet_name="Resumo", index=False, header=False)
        conciliados.to_excel(writer, sheet_name="Conciliados", index=False)
        divergencias.to_excel(writer, sheet_name="Divergencias", index=False)

        wb = writer.book
        _estilizar_planilha(wb["Conciliados"], conciliados)
        _estilizar_planilha(wb["Divergencias"], divergencias, cor_cabecalho="A8342A")

        ws_resumo = wb["Resumo"]
        ws_resumo.column_dimensions["A"].width = 42
        ws_resumo.column_dimensions["B"].width = 20
        ws_resumo["A1"].font = Font(name="Calibri", size=16, bold=True, color="2F4B3C")
        ws_resumo["A3"].font = Font(name="Calibri", size=9, color="5B6660")
        for ref in ("A5", "B5"):
            ws_resumo[ref].font = Font(name="Calibri", bold=True, color="FFFFFF")
            ws_resumo[ref].fill = PatternFill(start_color="2F4B3C", end_color="2F4B3C", fill_type="solid")

    return output.getvalue()


# =====================================================================
# HELPERS DE INTERFACE
# =====================================================================
def formatar_moeda(valor):
    if pd.isna(valor):
        return "—"
    texto = f"{valor:,.2f}"
    return "R$ " + texto.replace(",", "X").replace(".", ",").replace("X", ".")


def cartao_metrica(label, valor, sub="", tom="neutro"):
    st.markdown(
        f"""<div class="cb-card cb-card--{tom}">
            <div class="cb-card-label">{label}</div>
            <div class="cb-card-value">{valor}</div>
            <div class="cb-card-sub">{sub}</div>
        </div>""",
        unsafe_allow_html=True,
    )


def selo_percentual(pct, total_ok, total_geral):
    pct = max(0.0, min(100.0, pct))
    st.markdown(
        f"""<div class="cb-selo-wrap">
            <div class="cb-selo" style="background: conic-gradient(#2F4B3C 0% {pct}%, #E2DFD2 {pct}% 100%);">
                <div class="cb-selo-inner">
                    <span class="cb-selo-pct">{pct:.1f}%</span>
                    <span class="cb-selo-label">Conciliado</span>
                </div>
            </div>
            <div>
                <div class="cb-selo-titulo">Índice de conciliação</div>
                <div class="cb-selo-desc">{total_ok} de {total_geral} lançamentos batidos automaticamente.</div>
            </div>
        </div>""",
        unsafe_allow_html=True,
    )


def colorir_status(val):
    val = str(val)
    if "Banco" in val:
        return "background-color: #F7E8E6; color: #7A251D; font-weight: 600;"
    if "Sistema" in val:
        return "background-color: #F5EBDD; color: #6B4A22; font-weight: 600;"
    return ""


def colorir_tipo_match(val):
    if str(val) == "Exato":
        return "background-color: #E7EDE6; color: #2F4B3C; font-weight: 600;"
    return "background-color: #F5EBDD; color: #6B4A22; font-weight: 600;"


def ler_planilha(arquivo_upload, rotulo, key_prefix):
    arquivo_upload.seek(0)
    xls = pd.ExcelFile(arquivo_upload)
    aba = xls.sheet_names[0]
    if len(xls.sheet_names) > 1:
        aba = st.selectbox(f"Aba da planilha — {rotulo}", xls.sheet_names, key=f"{key_prefix}_aba")
    arquivo_upload.seek(0)
    df = pd.read_excel(arquivo_upload, sheet_name=aba)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def escolher_coluna(df, campo, palavras_chave, key):
    candidatas = detectar_colunas_candidatas(df.columns.tolist(), palavras_chave)
    opcoes = df.columns.tolist()
    default_idx = opcoes.index(candidatas[0]) if len(candidatas) == 1 else 0
    return st.selectbox(f"Coluna de {campo}", opcoes, index=default_idx, key=key)


# =====================================================================
# APP PRINCIPAL
# =====================================================================
def main():
    injetar_estilo()

    if "autenticado" not in st.session_state:
        st.session_state.autenticado = False
    if not st.session_state.autenticado:
        tela_login()
        return

    with st.sidebar:
        st.markdown(f"**👤 {st.session_state.usuario}**")
        if st.button("Sair", use_container_width=True):
            for chave in ("autenticado", "usuario", "resultado"):
                st.session_state.pop(chave, None)
            st.rerun()

        st.markdown('<hr class="cb-divisor" style="border-top-color:rgba(255,255,255,0.25);">', unsafe_allow_html=True)
        st.markdown("**⚙️ Configurações**")
        nome_empresa = st.text_input("Nome da empresa (aparece no relatório)", value="")
        tolerancia_dias = st.number_input(
            "Tolerância de dias entre lançamentos", min_value=0, max_value=15, value=0, step=1,
            help="Considera lançamentos como 'aproximados' quando o valor bate mas a data diverge em até N dias — útil para compensação bancária."
        )
        usar_abs = st.checkbox("Comparar por valor absoluto", value=False, help="Ative se um dos lados registra débitos como valor negativo e o outro como positivo.")

        st.markdown('<hr class="cb-divisor" style="border-top-color:rgba(255,255,255,0.25);">', unsafe_allow_html=True)
        if st.button("🔄 Nova conciliação", use_container_width=True):
            st.session_state.pop("resultado", None)
            st.rerun()

    st.markdown(
        f"""<div class="cb-header">
            <div class="cb-header-eyebrow">Fechamento contábil</div>
            <h1>📒 {NOME_PRODUTO}</h1>
            <p>Cruze o extrato bancário com o sistema, veja exatamente onde estão as diferenças
            e exporte um relatório pronto para o cliente — em segundos.</p>
        </div>""",
        unsafe_allow_html=True,
    )

    col_up1, col_up2 = st.columns(2)
    with col_up1:
        arquivo_banco = st.file_uploader("1. Extrato Bancário (.xlsx)", type=["xlsx"])
    with col_up2:
        arquivo_sistema = st.file_uploader("2. Relatório do Sistema (.xlsx)", type=["xlsx"])

    if not (arquivo_banco and arquivo_sistema):
        st.info("🔒 Envie as duas planilhas para continuar.")
        return

    try:
        df_banco_raw = ler_planilha(arquivo_banco, "Extrato Bancário", "banco")
        df_sistema_raw = ler_planilha(arquivo_sistema, "Sistema", "sistema")
    except Exception as e:
        st.error(f"❌ Não consegui abrir uma das planilhas. Verifique se é um .xlsx válido. Detalhe técnico: {e}")
        return

    if df_banco_raw.empty or df_sistema_raw.empty:
        st.error("❌ Uma das planilhas está vazia.")
        return

    cand_data_b = detectar_colunas_candidatas(df_banco_raw.columns, ["data", "dt"])
    cand_valor_b = detectar_colunas_candidatas(df_banco_raw.columns, ["val"])
    cand_data_s = detectar_colunas_candidatas(df_sistema_raw.columns, ["data", "dt"])
    cand_valor_s = detectar_colunas_candidatas(df_sistema_raw.columns, ["val"])
    precisa_mapeamento = not all(len(c) == 1 for c in (cand_data_b, cand_valor_b, cand_data_s, cand_valor_s))

    with st.expander("🔧 Mapeamento de colunas", expanded=precisa_mapeamento):
        st.caption("Confirme quais colunas representam Data e Valor em cada planilha.")
        col_map1, col_map2 = st.columns(2)
        with col_map1:
            st.markdown("**Extrato Bancário**")
            col_data_banco = escolher_coluna(df_banco_raw, "Data", ["data", "dt"], "map_data_banco")
            col_valor_banco = escolher_coluna(df_banco_raw, "Valor", ["val"], "map_valor_banco")
        with col_map2:
            st.markdown("**Sistema**")
            col_data_sistema = escolher_coluna(df_sistema_raw, "Data", ["data", "dt"], "map_data_sistema")
            col_valor_sistema = escolher_coluna(df_sistema_raw, "Valor", ["val"], "map_valor_sistema")

    if st.button("🚀 Processar Conciliação", use_container_width=True, type="primary"):
        with st.spinner("Cruzando os lançamentos..."):
            try:
                df_banco, invalidos_banco = preparar_dataframe(df_banco_raw, col_data_banco, col_valor_banco, usar_abs)
                df_sistema, invalidos_sistema = preparar_dataframe(df_sistema_raw, col_data_sistema, col_valor_sistema, usar_abs)

                if df_banco.empty or df_sistema.empty:
                    st.error("❌ Depois de limpar os dados, uma das planilhas ficou sem linhas válidas de Data/Valor. Confira o mapeamento de colunas acima.")
                    return

                conciliados, divergencias = executar_conciliacao(df_banco, df_sistema, tolerancia_dias)

                total_banco = len(df_banco)
                total_sistema = len(df_sistema)
                total_conciliados = len(conciliados)
                total_divergencias = len(divergencias)
                base_pct = max(total_banco, total_sistema, 1)
                percentual = 100 * total_conciliados / base_pct

                resumo = {
                    "total_banco": total_banco, "total_sistema": total_sistema,
                    "total_conciliados": total_conciliados, "total_divergencias": total_divergencias,
                    "percentual": percentual,
                }
                excel_bytes = gerar_excel_conciliacao(conciliados, divergencias, resumo, nome_empresa)

                st.session_state.resultado = {
                    "conciliados": conciliados, "divergencias": divergencias, "resumo": resumo,
                    "excel_bytes": excel_bytes, "invalidos_banco": invalidos_banco, "invalidos_sistema": invalidos_sistema,
                }
            except Exception as e:
                st.error(f"❌ Erro ao processar: {e}")
                return

    if "resultado" not in st.session_state:
        return

    r = st.session_state.resultado
    resumo = r["resumo"]

    if r["invalidos_banco"] or r["invalidos_sistema"]:
        st.warning(
            f"⚠️ Foram ignoradas {r['invalidos_banco']} linha(s) do Extrato e "
            f"{r['invalidos_sistema']} linha(s) do Sistema por terem Data ou Valor inválidos/vazios."
        )

    st.markdown('<hr class="cb-divisor">', unsafe_allow_html=True)
    st.subheader("📊 Resumo")

    col_a, col_b = st.columns([1.1, 2])
    with col_a:
        selo_percentual(resumo["percentual"], resumo["total_conciliados"], max(resumo["total_banco"], resumo["total_sistema"]))
    with col_b:
        c1, c2, c3 = st.columns(3)
        with c1:
            cartao_metrica("Extrato Bancário", resumo["total_banco"], "lançamentos", "neutro")
        with c2:
            cartao_metrica("Conciliados", resumo["total_conciliados"], "batidos automaticamente", "verde")
        with c3:
            cartao_metrica("Divergências", resumo["total_divergencias"], "precisam de revisão", "vermelho" if resumo["total_divergencias"] else "neutro")

    st.markdown('<hr class="cb-divisor">', unsafe_allow_html=True)

    aba_conc, aba_div = st.tabs([f"✅ Conciliados ({len(r['conciliados'])})", f"⚠️ Divergências ({len(r['divergencias'])})"])

    with aba_conc:
        if r["conciliados"].empty:
            st.caption("Nenhum lançamento conciliado ainda.")
        else:
            df_show = r["conciliados"].copy()
            for c in ("Valor_Banco", "Valor_Sistema"):
                if c in df_show.columns:
                    df_show[c] = df_show[c].apply(formatar_moeda)
            for c in ("Data_Banco", "Data_Sistema"):
                if c in df_show.columns:
                    df_show[c] = pd.to_datetime(df_show[c]).dt.strftime("%d/%m/%Y")
            estilo = df_show.style.map(colorir_tipo_match, subset=["Tipo_de_Match"]) if "Tipo_de_Match" in df_show.columns else df_show.style
            st.dataframe(estilo, use_container_width=True, hide_index=True)

    with aba_div:
        if r["divergencias"].empty:
            st.caption("Nenhuma divergência — tudo bateu! 🎉")
        else:
            df_show = r["divergencias"].copy()
            for c in ("Valor_Banco", "Valor_Sistema"):
                if c in df_show.columns:
                    df_show[c] = df_show[c].apply(formatar_moeda)
            for c in ("Data_Banco", "Data_Sistema"):
                if c in df_show.columns:
                    df_show[c] = pd.to_datetime(df_show[c]).dt.strftime("%d/%m/%Y")
            estilo = df_show.style.map(colorir_status, subset=["Status"]) if "Status" in df_show.columns else df_show.style
            st.dataframe(estilo, use_container_width=True, hide_index=True)

    st.markdown('<hr class="cb-divisor">', unsafe_allow_html=True)
    st.download_button(
        label="📥 Baixar Relatório Completo (.xlsx)",
        data=r["excel_bytes"],
        file_name=f"conciliacao_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
